from django.core.paginator import Paginator, PageNotAnInteger, EmptyPage
from django.db import models
from django.http import HttpResponse
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth import authenticate, login, logout
from django.contrib.auth.decorators import login_required
from django.contrib.admin.views.decorators import staff_member_required
from django.contrib import messages
from django.db import IntegrityError
from django.contrib.auth.models import User
from django.utils import timezone
from django.contrib.sessions.models import Session
from .models import Question, TestPaper, Profile, TestRecord, AnswerRecord, WrongQuestion, Class, ClassAdmin, ClassApplication, ClassAssignment, ClassAssignmentRecord, Subject, Chapter, Section, KnowledgePoint
from .utils import paginate_queryset, compare_answers, calculate_score, parse_datetime_local, download_template_response, import_questions_from_excel
from .captcha import generate_captcha_text, generate_captcha_image
import datetime
import json
import re
import hashlib
from django.core.cache import cache

# 导入文件缓存（内存存储，重启后失效但避免重复导入）
_imported_files_cache = {}

def generate_file_hash(file_content):
    return hashlib.md5(file_content).hexdigest()

def is_duplicate_import(file_content, window_hours=24):
    file_hash = generate_file_hash(file_content)
    if file_hash in _imported_files_cache:
        import_time, imported_count = _imported_files_cache[file_hash]
        time_diff = timezone.now() - import_time
        if time_diff.total_seconds() < window_hours * 3600:
            return True, imported_count, import_time
    return False, 0, None

def get_visible_questions(user):
    """获取用户可见的题目：公开题目 + 用户自己创建的私有题目"""
    if user.is_staff:
        return Question.objects.all()
    return Question.objects.filter(
        models.Q(is_public=True) | models.Q(created_by=user.username)
    )

# 答题视图
def question_detail(request, question_id):
    question = get_object_or_404(Question, pk=question_id)
    if request.method == 'POST':
        user_answer = request.POST.get('answer')
        if user_answer == question.correct_answer:
            result = '正确'
        else:
            result = '错误'
        return render(request, 'quiz/frontend/answer_result.html', {
            'question': question,
            'user_answer': user_answer,
            'result': result,
            'correct_answer': question.correct_answer
        })
    return render(request, 'quiz/frontend/question_detail.html', {'question': question})

def test_paper_list(request):
    user = request.user
    if user.is_staff:
        test_papers = TestPaper.objects.filter(is_published=True).order_by('-created_at')
    else:
        test_papers = TestPaper.objects.filter(
            is_published=True
        ).filter(
            models.Q(is_public=True) | models.Q(created_by=user.username)
        ).order_by('-created_at')
    paginated_test_papers = paginate_queryset(test_papers, request.GET.get('page', 1))
    return render(request, 'quiz/frontend/test_paper_list.html', {
        'test_papers': paginated_test_papers
    })

def test_paper_detail(request, paper_id):
    test_paper = get_object_or_404(TestPaper, pk=paper_id)
    user = request.user
    if not test_paper.is_public and (not user.is_authenticated or (test_paper.created_by != user.username and not user.is_staff)):
        raise Http404('试卷不存在或无权访问')
    questions = list(test_paper.questions.all())
    for q in questions:
        if isinstance(q.options, str):
            import json
            try:
                q.options = json.loads(q.options)
            except:
                q.options = {}
    return render(request, 'quiz/frontend/test_paper_detail.html', {
        'test_paper': test_paper,
        'questions': questions
    })

@login_required
def submit_test_paper(request, paper_id):
    test_paper = get_object_or_404(TestPaper, pk=paper_id)
    questions = list(test_paper.questions.all())
    
    if request.method == 'POST':
        user_answers = {}
        for q in questions:
            answer_key = f'question_{q.id}'
            if answer_key in request.POST:
                user_answers[q.id] = request.POST[answer_key]
        
        score, correct_count, wrong_count, total_count, question_results = calculate_score(questions, user_answers)
        
        # 查找或创建用户档案
        try:
            profile = Profile.objects.get(user=request.user)
        except Profile.DoesNotExist:
            profile = Profile.objects.create(user=request.user)
        
        # 创建答题记录
        test_record = TestRecord.objects.create(
            user=request.user,
            test_paper=test_paper,
            score=score,
            total_score=test_paper.total_score,
            completed_at=timezone.now()
        )
        
        # 创建答案记录并收集错题
        wrong_questions_list = []
        for result in question_results:
            AnswerRecord.objects.create(
                test_record=test_record,
                question=result['question'],
                user_answer=result.get('user_answer', ''),
                correct_answer=result['correct_answer'],
                is_correct=result['is_correct'],
                score=result['score']
            )
            
            if not result['is_correct']:
                wrong_questions_list.append(result['question'])
        
        # 自动添加错题到错题本
        for question in wrong_questions_list:
            WrongQuestion.objects.get_or_create(
                user=request.user,
                question=question,
                defaults={
                    'user_answer': user_answers.get(question.id, ''),
                    'correct_answer': question.correct_answer
                }
            )
        
        # 更新用户学习统计
        profile.total_score += score
        profile.tests_taken += 1
        profile.accuracy_rate = round(correct_count / total_count * 100, 1) if total_count > 0 else 0
        profile.save()
        
        return render(request, 'quiz/frontend/test_paper_result.html', {
            'test_paper': test_paper,
            'score': score,
            'correct_count': correct_count,
            'wrong_count': wrong_count,
            'total_count': total_count,
            'question_results': question_results,
            'test_record': test_record
        })
    
    for q in questions:
        if isinstance(q.options, str):
            import json
            try:
                q.options = json.loads(q.options)
            except:
                q.options = {}
    
    return render(request, 'quiz/frontend/test_paper_detail.html', {
        'test_paper': test_paper,
        'questions': questions
    })

@login_required
def test_history(request):
    test_records = TestRecord.objects.filter(user=request.user).order_by('-completed_at')
    paginated_records = paginate_queryset(test_records, request.GET.get('page', 1), items_per_page=10)
    
    for record in paginated_records:
        if isinstance(record.test_paper.questions, list):
            record.question_count = len(record.test_paper.questions)
        else:
            record.question_count = record.test_paper.questions.count()
    
    return render(request, 'quiz/frontend/test_history.html', {
        'test_records': paginated_records
    })

@login_required
def test_history_detail(request, record_id):
    test_record = get_object_or_404(TestRecord, pk=record_id, user=request.user)
    answer_records = AnswerRecord.objects.filter(test_record=test_record).select_related('question')
    
    question_results = []
    for ar in answer_records:
        question_results.append({
            'question': ar.question,
            'user_answer': ar.user_answer,
            'correct_answer': ar.correct_answer,
            'is_correct': ar.is_correct,
            'score': ar.score
        })
    
    return render(request, 'quiz/frontend/test_history_detail.html', {
        'test_record': test_record,
        'question_results': question_results
    })

def login_view(request):
    if request.user.is_authenticated:
        return redirect('user_center')
    
    if request.method == 'POST':
        captcha = request.POST.get('captcha', '').strip()
        expected_captcha = cache.get(f'captcha:{request.session.session_key}', '')
        
        if not captcha or captcha.lower() != expected_captcha.lower():
            messages.error(request, '验证码错误，请重试')
            return render(request, 'quiz/frontend/login.html')
        
        username = request.POST.get('username', '').strip()
        password = request.POST.get('password', '')
        
        if not username or not password:
            messages.error(request, '请输入用户名和密码')
            return render(request, 'quiz/frontend/login.html')
        
        user = authenticate(username=username, password=password)
        
        if user is None:
            try:
                user_obj = User.objects.get(username=username)
                if not user_obj.is_active:
                    messages.error(request, '用户被停用，请联系管理员激活')
                    return render(request, 'quiz/frontend/login.html')
            except User.DoesNotExist:
                pass
            
            try:
                profile = Profile.objects.get(phone_number=username)
                user = authenticate(username=profile.user.username, password=password)
                if user is None:
                    messages.error(request, '用户名/手机号码或密码错误')
                    return render(request, 'quiz/frontend/login.html')
                if not user.is_active:
                    messages.error(request, '用户被停用，请联系管理员激活')
                    return render(request, 'quiz/frontend/login.html')
            except Profile.DoesNotExist:
                messages.error(request, '用户名/手机号码或密码错误')
                return render(request, 'quiz/frontend/login.html')
        
        if user is not None:
            login(request, user)

            try:
                profile = Profile.objects.get(user=user)
                if profile.approval_status == 0:
                    logout(request)
                    return redirect('approval_pending')
                elif profile.approval_status == 2:
                    logout(request)
                    messages.error(request, '您的账号已被拒绝，请联系管理员')
                    return render(request, 'quiz/frontend/login.html')

                if profile.session_key and profile.session_key != request.session.session_key:
                    try:
                        old_session = Session.objects.get(session_key=profile.session_key)
                        old_session.delete()
                    except Session.DoesNotExist:
                        pass

                profile.session_key = request.session.session_key
                profile.save(update_fields=['session_key'])
            except Profile.DoesNotExist:
                pass

            next_url = request.GET.get('next', 'user_center')
            return redirect(next_url)
        else:
            messages.error(request, '用户名/手机号码或密码错误')
    
    captcha_text = generate_captcha_text()
    captcha_image = generate_captcha_image(captcha_text)
    cache.set(f'captcha:{request.session.session_key}', captcha_text, 300)
    
    return render(request, 'quiz/frontend/login.html', {
        'captcha_image': captcha_image
    })

def register(request):
    if request.user.is_authenticated:
        return redirect('user_center')
    
    if request.method == 'POST':
        captcha = request.POST.get('captcha', '').strip()
        expected_captcha = cache.get(f'captcha:{request.session.session_key}', '')
        
        if not captcha or captcha.lower() != expected_captcha.lower():
            messages.error(request, '验证码错误，请重试')
            return redirect('register')
        
        username = request.POST.get('username', '').strip()
        email = request.POST.get('email', '').strip()
        password = request.POST.get('password', '')
        confirm_password = request.POST.get('confirm_password', '')
        first_name = request.POST.get('first_name', '').strip()
        phone_number = request.POST.get('phone_number', '').strip()
        qq_number = request.POST.get('qq_number', '').strip()
        
        if not username or not email or not password or not first_name or not phone_number:
            messages.error(request, '用户名、姓名、邮箱、手机号码和密码为必填项')
            return render(request, 'quiz/frontend/register.html')
        
        if password != confirm_password:
            messages.error(request, '两次输入的密码不一致')
            return render(request, 'quiz/frontend/register.html')
        
        if len(password) < 6:
            messages.error(request, '密码长度不能少于6位')
            return render(request, 'quiz/frontend/register.html')
        
        phone_pattern = re.compile(r'^1[3-9]\d{9}$')
        if not phone_pattern.match(phone_number):
            messages.error(request, '请输入有效的11位手机号码')
            return render(request, 'quiz/frontend/register.html')
        
        if User.objects.filter(username=username).exists():
            messages.error(request, '用户名已存在')
            return render(request, 'quiz/frontend/register.html')
        
        if User.objects.filter(email=email).exists():
            messages.error(request, '邮箱已被注册')
            return render(request, 'quiz/frontend/register.html')
        
        if Profile.objects.filter(phone_number=phone_number).exists():
            messages.error(request, '手机号码已被注册')
            return render(request, 'quiz/frontend/register.html')
        
        try:
            user = User.objects.create_user(
                username=username,
                email=email,
                password=password,
                first_name=first_name
            )
            
            Profile.objects.create(
                user=user,
                phone_number=phone_number,
                qq_number=qq_number,
                approval_status=0
            )
            
            messages.success(request, '注册成功！您的账号正在等待管理员审核通过。')
            return redirect('login')
        except Exception as e:
            messages.error(request, f'注册失败：{str(e)}')
    
    return render(request, 'quiz/frontend/register.html')

def approval_pending(request):
    return render(request, 'quiz/frontend/approval_pending.html')

def captcha_image(request):
    captcha_text = generate_captcha_text()
    captcha_image = generate_captcha_image(captcha_text)
    cache.set(f'captcha:{request.session.session_key}', captcha_text, 300)
    return HttpResponse(captcha_image, content_type='image/png')

def refresh_captcha(request):
    return captcha_image(request)

def logout_view(request):
    logout(request)
    return redirect('login')

@login_required
def user_center(request):
    try:
        profile = Profile.objects.get(user=request.user)
    except Profile.DoesNotExist:
        profile = Profile.objects.create(user=request.user)
    
    recent_tests = TestRecord.objects.filter(user=request.user).order_by('-completed_at')[:5]
    recent_wrong_questions = WrongQuestion.objects.filter(user=request.user).order_by('-added_at')[:5]
    
    context = {
        'profile': profile,
        'recent_tests': recent_tests,
        'recent_wrong_questions': recent_wrong_questions,
        'is_admin': request.user.is_staff
    }
    
    return render(request, 'quiz/frontend/user_center.html', context)

@login_required
def wrong_question_notebook(request):
    wrong_questions = WrongQuestion.objects.filter(user=request.user).select_related('question').order_by('-added_at')
    
    wrong_questions_list = []
    for wq in wrong_questions:
        q = wq.question
        if isinstance(q.options, str):
            import json
            try:
                q.options = json.loads(q.options)
            except:
                q.options = {}
        wrong_questions_list.append({
            'wrong_question': wq,
            'question': q
        })
    
    paginated_wrong_questions = paginate_queryset(wrong_questions_list, request.GET.get('page', 1), items_per_page=10)
    
    return render(request, 'quiz/frontend/wrong_question_notebook.html', {
        'wrong_questions': paginated_wrong_questions
    })

@login_required
def create_wrong_question_paper(request):
    wrong_questions = WrongQuestion.objects.filter(user=request.user).select_related('question')
    
    questions = []
    for wq in wrong_questions:
        q = wq.question
        if isinstance(q.options, str):
            import json
            try:
                q.options = json.loads(q.options)
            except:
                q.options = {}
        questions.append(q)
    
    if not questions:
        messages.error(request, '您的错题本中没有题目')
        return redirect('wrong_question_notebook')
    
    if request.method == 'POST':
        title = request.POST.get('title', '错题巩固试卷')
        
        test_paper = TestPaper.objects.create(
            title=title,
            description='错题巩固试卷',
            created_by=request.user.username,
            is_published=False
        )
        
        for q in questions:
            test_paper.questions.add(q)
        
        test_paper.total_score = sum(q.score for q in questions)
        test_paper.save()
        
        return redirect('submit_wrong_question_paper', paper_id=test_paper.id)
    
    return render(request, 'quiz/frontend/wrong_question_paper.html', {
        'questions': questions
    })

@login_required
def submit_wrong_question_paper(request, paper_id):
    test_paper = get_object_or_404(TestPaper, pk=paper_id)
    questions = list(test_paper.questions.all())
    
    if request.method == 'POST':
        user_answers = {}
        for q in questions:
            answer_key = f'question_{q.id}'
            if answer_key in request.POST:
                user_answers[q.id] = request.POST[answer_key]
        
        score, correct_count, wrong_count, total_count, question_results = calculate_score(questions, user_answers)
        
        test_record = TestRecord.objects.create(
            user=request.user,
            test_paper=test_paper,
            score=score,
            total_score=test_paper.total_score,
            completed_at=timezone.now(),
            is_wrong_paper=True
        )
        
        wrong_question_ids = set()
        for result in question_results:
            AnswerRecord.objects.create(
                test_record=test_record,
                question=result['question'],
                user_answer=result.get('user_answer', ''),
                correct_answer=result['correct_answer'],
                is_correct=result['is_correct'],
                score=result['score']
            )
        
        for wq in WrongQuestion.objects.filter(user=request.user, question__in=questions):
            wrong_question_ids.add(wq.question.id)
        
        WrongQuestion.objects.filter(user=request.user, question__in=questions).delete()
        
        return render(request, 'quiz/frontend/wrong_question_paper_result.html', {
            'test_paper': test_paper,
            'score': score,
            'correct_count': correct_count,
            'wrong_count': wrong_count,
            'total_count': total_count,
            'question_results': question_results,
            'test_record': test_record,
            'deleted_wrong_questions': len(wrong_question_ids)
        })
    
    for q in questions:
        if isinstance(q.options, str):
            import json
            try:
                q.options = json.loads(q.options)
            except:
                q.options = {}
    
    return render(request, 'quiz/frontend/test_paper_detail.html', {
        'test_paper': test_paper,
        'questions': questions,
        'is_wrong_paper': True
    })

@login_required
def delete_wrong_question(request, wrong_question_id):
    wrong_question = get_object_or_404(WrongQuestion, pk=wrong_question_id, user=request.user)
    wrong_question.delete()
    messages.success(request, '已从错题本中删除')
    return redirect('wrong_question_notebook')

@login_required
def my_test_papers(request):
    test_papers = TestPaper.objects.filter(created_by=request.user.username).order_by('-created_at')
    paginated_test_papers = paginate_queryset(test_papers, request.GET.get('page', 1))
    
    for paper in paginated_test_papers:
        paper.question_count = paper.questions.count()
    
    return render(request, 'quiz/frontend/my_test_papers.html', {
        'test_papers': paginated_test_papers
    })

@login_required
def create_test_paper(request):
    """创建试卷视图 - 支持手动添加题目"""
    if request.method == 'POST':
        title = request.POST.get('title')
        description = request.POST.get('description')
        question_ids = request.POST.getlist('questions')

        if title and question_ids:
            test_paper = TestPaper.objects.create(
                title=title,
                description=description,
                created_by=request.user.username,
                is_published=False
            )

            total_score = 0
            for q_id in question_ids:
                try:
                    question = Question.objects.get(id=q_id)
                    test_paper.questions.add(question)
                    total_score += question.score
                except Question.DoesNotExist:
                    pass

            test_paper.total_score = total_score
            test_paper.save()

            messages.success(request, f'试卷 "{title}" 创建成功！共 {len(question_ids)} 道题目，总分 {total_score} 分。')
            return redirect('my_test_papers')
        else:
            messages.error(request, '请填写试卷标题并至少选择一道题目')

    subjects = Subject.objects.all().order_by('name')
    chapters = Chapter.objects.select_related('subject').order_by('subject', 'number')
    knowledge_points = KnowledgePoint.objects.select_related('section', 'section__chapter', 'subject').order_by('subject', 'name')

    questions = get_visible_questions(request.user).select_related('subject', 'chapter', 'section').prefetch_related('knowledge_points').order_by('id')

    questions_list = []
    for q in questions:
        options_data = q.options
        if isinstance(options_data, str):
            import json
            try:
                options_data = json.loads(options_data)
            except:
                options_data = {}
        elif not isinstance(options_data, dict):
            options_data = {}
        
        kp_ids = [str(kp.id) for kp in q.knowledge_points.all()]
        
        questions_list.append({
            'id': q.id,
            'type': q.type,
            'content': q.content,
            'options': options_data,
            'score': q.score,
            'explanation': q.explanation,
            'subject_id': q.subject.id if q.subject else '',
            'chapter_id': q.chapter.id if q.chapter else '',
            'knowledge_point_ids': kp_ids
        })
    
    return render(request, 'quiz/frontend/create_test_paper.html', {
        'questions': questions_list,
        'subjects': subjects,
        'chapters': chapters,
        'knowledge_points': knowledge_points
    })

@login_required
def import_test_paper(request):
    """导入试卷视图 - 支持从Excel文件导入试卷"""
    
    if request.method == 'POST' and request.POST.get('action') == 'save_and_back':
        title = request.POST.get('title', '')
        description = request.POST.get('description', '')
        questions_json = request.POST.get('questions_json', '')
        
        request.session['import_title'] = title
        request.session['import_description'] = description
        request.session['import_questions_json'] = questions_json
        
        messages.success(request, '编辑内容已保存')
        return redirect('import_test_paper')
    
    if request.method == 'POST' and request.POST.get('action') == 'preview':
        file = request.FILES.get('file')
        if not file:
            messages.error(request, '请上传文件')
            return render(request, 'quiz/frontend/import_test_paper.html')
        
        try:
            questions_data, stats, errors = import_questions_from_excel(file)
            if errors:
                messages.error(request, errors[0])
                return render(request, 'quiz/frontend/import_test_paper.html')
            
            for idx, q in enumerate(questions_data):
                q['row'] = idx + 2
                q['has_error'] = not (q.get('correct_answer') and q.get('score'))
            
            return render(request, 'quiz/frontend/import_preview.html', {
                'questions_data': questions_data,
                'questions_json': json.dumps(questions_data, ensure_ascii=False),
                'total_score': sum(q.get('score', 0) for q in questions_data if isinstance(q.get('score'), int)),
                'valid_count': sum(1 for q in questions_data if q.get('correct_answer') and q.get('score')),
                'missing_count': sum(1 for q in questions_data if not q.get('correct_answer') or not q.get('score'))
            })
        except Exception as e:
            messages.error(request, f'预览失败：{str(e)}')
    
    if request.method == 'POST' and request.POST.get('action') == 'confirm_import':
        title = request.POST.get('title')
        description = request.POST.get('description')
        questions_json = request.POST.get('questions_data')
        
        if title and questions_json:
            try:
                questions_data = json.loads(questions_json)
                
                if not questions_data:
                    messages.error(request, '没有有效的题目数据')
                    return render(request, 'quiz/frontend/import_test_paper.html')
                
                test_paper = TestPaper.objects.create(
                    title=title,
                    description=description,
                    created_by=request.user.username,
                    is_published=False
                )
                
                total_score = 0
                valid_questions = 0
                for q_data in questions_data:
                    if not q_data.get('content') or not q_data.get('correct_answer'):
                        continue
                    
                    options_data = q_data.get('options', {})
                    if isinstance(options_data, str):
                        try:
                            options_data = json.loads(options_data)
                        except:
                            options_data = {}
                    
                    q_type = int(q_data.get('type', 1))
                    if q_type not in [1, 2, 3]:
                        q_type = 1
                    
                    q_score = q_data.get('score', 1)
                    try:
                        q_score = int(q_score) if q_score else 1
                    except:
                        q_score = 1
                    
                    subject_obj = None
                    chapter_obj = None
                    section_obj = None
                    kp_objects = []
                    
                    subject_name = q_data.get('subject_name', '').strip()
                    chapter_title = q_data.get('chapter_title', '').strip()
                    section_title = q_data.get('section_title', '').strip()
                    kp_names = q_data.get('knowledge_points_str', '').strip()
                    
                    if subject_name:
                        subject_obj, _ = Subject.objects.get_or_create(
                            name=subject_name,
                            defaults={'code': subject_name[:10].upper(), 'icon': '📚'}
                        )
                    
                    if subject_obj and chapter_title:
                        chapter_obj, _ = Chapter.objects.get_or_create(
                            subject=subject_obj,
                            title=chapter_title,
                            defaults={'number': Chapter.objects.filter(subject=subject_obj).count() + 1}
                        )
                    
                    if chapter_obj and section_title:
                        section_obj, _ = Section.objects.get_or_create(
                            chapter=chapter_obj,
                            title=section_title,
                            defaults={'number': Section.objects.filter(chapter=chapter_obj).count() + 1}
                        )
                    
                    if kp_names and subject_obj:
                        for kp_name in kp_names.split(','):
                            kp_name = kp_name.strip()
                            if kp_name:
                                kp, created = KnowledgePoint.objects.get_or_create(
                                    subject=subject_obj,
                                    name=kp_name,
                                    defaults={'section': section_obj}
                                )
                                kp_objects.append(kp)
                    
                    question = Question.objects.create(
                        type=q_type,
                        content=q_data['content'],
                        options=options_data,
                        correct_answer=q_data['correct_answer'],
                        score=q_score,
                        explanation=q_data.get('explanation', ''),
                        subject=subject_obj,
                        chapter=chapter_obj,
                        section=section_obj,
                        is_public=False,
                        created_by=request.user.username
                    )
                    
                    if kp_objects:
                        question.knowledge_points.set(kp_objects)
                    
                    test_paper.questions.add(question)
                    total_score += q_score
                    valid_questions += 1
                
                test_paper.total_score = total_score
                test_paper.save()
                
                messages.success(request, f'试卷 "{title}" 导入成功！共导入 {len(questions_data)} 道题目')
                return redirect('my_test_papers')
            except Exception as e:
                messages.error(request, f'导入失败：{str(e)}')
        else:
            messages.error(request, '请填写试卷标题并确认导入')
        return render(request, 'quiz/frontend/import_test_paper.html')
    
    if request.method == 'POST' and request.FILES.get('paper_file'):
        title = request.POST.get('title', '')
        description = request.POST.get('description', '')
        file = request.FILES.get('paper_file')
        
        if not title:
            messages.error(request, '请填写试卷标题')
            return render(request, 'quiz/frontend/import_test_paper.html')
        
        try:
            import openpyxl
            from openpyxl.utils.exceptions import InvalidFileException
            
            wb = openpyxl.load_workbook(file)
            ws = wb.active
            
            headers = [cell.value for cell in ws[1]]
            questions_data = []
            
            for row_idx, row in enumerate(ws.iter_rows(min_row=2), start=2):
                row_data = {headers[i]: row[i].value for i in range(len(headers))}
                content = row_data.get('题目内容', '')
                if not content:
                    continue
                
                q_type_map = {'单选题': 1, '多选题': 2, '判断题': 3, '1': 1, '2': 2, '3': 3}
                q_type = q_type_map.get(str(row_data.get('题型', '单选题')), 1)
                
                options = {}
                for letter in ['A', 'B', 'C', 'D']:
                    option_key = f'选项{letter}'
                    if row_data.get(option_key):
                        options[letter] = row_data[option_key]
                
                correct_answer = str(row_data.get('正确答案', '')).strip()
                score = row_data.get('分值', 1)
                try:
                    score = int(score) if score else 1
                except:
                    score = 1
                
                subject_name = str(row_data.get('学科', '') or '').strip()
                chapter_title = str(row_data.get('章节', '') or '').strip()
                section_title = str(row_data.get('小节', '') or '').strip()
                kp_names = str(row_data.get('知识点', '') or '').strip()
                
                questions_data.append({
                    'content': content,
                    'type': q_type,
                    'options': options,
                    'correct_answer': correct_answer,
                    'score': score,
                    'explanation': row_data.get('解析', ''),
                    'subject_name': subject_name,
                    'chapter_title': chapter_title,
                    'section_title': section_title,
                    'knowledge_points_str': kp_names,
                    'row': row_idx,
                    'has_error': not (correct_answer and score)
                })
            
            if not questions_data:
                messages.error(request, '文件中没有有效的题目数据')
                return render(request, 'quiz/frontend/import_test_paper.html')
            
            valid_count = sum(1 for q in questions_data if not q['has_error'])
            missing_count = len(questions_data) - valid_count
            
            return render(request, 'quiz/frontend/import_preview.html', {
                'questions_data': questions_data,
                'questions_json': json.dumps(questions_data, ensure_ascii=False),
                'total_score': sum(q['score'] for q in questions_data),
                'valid_count': valid_count,
                'missing_count': missing_count,
                'title': title,
                'description': description
            })
        except InvalidFileException:
            messages.error(request, '文件格式不正确，请上传 .xlsx 格式的 Excel 文件')
        except Exception as e:
            messages.error(request, f'读取文件失败：{str(e)}')
    
    saved_title = request.session.get('import_title', '')
    saved_description = request.session.get('import_description', '')
    
    return render(request, 'quiz/frontend/import_test_paper.html', {
        'saved_title': saved_title,
        'saved_description': saved_description
    })

@login_required
def publish_test_paper(request, paper_id):
    test_paper = get_object_or_404(TestPaper, pk=paper_id, created_by=request.user.username)
    
    if request.method == 'POST':
        test_paper.is_published = True
        test_paper.save()
        messages.success(request, f'试卷 "{test_paper.title}" 已发布到全站')
        return redirect('my_test_papers')
    
    return render(request, 'quiz/frontend/publish_test_paper.html', {
        'test_paper': test_paper
    })

@login_required
def delete_test_paper(request, paper_id):
    test_paper = get_object_or_404(TestPaper, pk=paper_id, created_by=request.user.username)
    
    if request.method == 'POST':
        test_paper.delete()
        messages.success(request, '试卷已删除')
        return redirect('my_test_papers')
    
    return render(request, 'quiz/frontend/delete_test_paper.html', {
        'test_paper': test_paper
    })

@login_required
def class_list(request):
    user_classes = Class.objects.filter(
        class_admins__user=request.user
    ).distinct().order_by('-created_at')
    
    student_classes = Class.objects.filter(
        profiles__user=request.user
    ).exclude(
        id__in=user_classes.values_list('id', flat=True)
    ).distinct()
    
    all_classes = (list(user_classes) + list(student_classes))
    unique_classes = []
    seen_ids = set()
    for cls in all_classes:
        if cls.id not in seen_ids:
            unique_classes.append(cls)
            seen_ids.add(cls.id)
    
    return render(request, 'quiz/frontend/class_list.html', {
        'classes': unique_classes
    })

@login_required
def class_detail(request, class_id):
    class_obj = get_object_or_404(Class, pk=class_id)
    
    is_admin = ClassAdmin.objects.filter(class_obj=class_obj, user=request.user).exists()
    
    if not is_admin:
        is_student = Class.objects.filter(id=class_id, profiles__user=request.user).exists()
        if not is_student:
            messages.error(request, '您不是该班级的管理员或学生')
            return redirect('class_list')
    
    admins = ClassAdmin.objects.filter(class_obj=class_obj).select_related('user', 'user__profile')
    students = Profile.objects.filter(class_obj=class_obj, approval_status=1).select_related('user', 'user__profile')
    pending_applications = ClassApplication.objects.filter(class_obj=class_obj, status=0).select_related('user')
    
    return render(request, 'quiz/frontend/class_detail.html', {
        'class_obj': class_obj,
        'admins': admins,
        'students': students,
        'pending_applications': pending_applications,
        'pending_count': pending_applications.count(),
        'is_admin': is_admin
    })

@login_required
def create_class(request):
    if request.method == 'POST':
        name = request.POST.get('name', '').strip()
        description = request.POST.get('description', '').strip()
        code = request.POST.get('code', '').strip()
        join_rule = request.POST.get('join_rule', 'approval').strip()
        
        if not name or not code:
            messages.error(request, '请填写班级名称和班级代码')
            return render(request, 'quiz/frontend/create_class.html')
        
        if Class.objects.filter(code=code).exists():
            messages.error(request, '班级代码已存在')
            return render(request, 'quiz/frontend/create_class.html')
        
        class_obj = Class.objects.create(
            name=name,
            description=description,
            code=code,
            join_rule=join_rule
        )
        
        ClassAdmin.objects.create(
            class_obj=class_obj,
            user=request.user
        )
        
        try:
            profile = Profile.objects.get(user=request.user)
            profile.class_obj = class_obj
            profile.save()
        except Profile.DoesNotExist:
            Profile.objects.create(user=request.user, class_obj=class_obj)
        
        messages.success(request, f'班级 "{name}" 创建成功！')
        return redirect('class_detail', class_id=class_obj.id)
    
    return render(request, 'quiz/frontend/create_class.html')

@login_required
def edit_class(request, class_id):
    class_obj = get_object_or_404(Class, pk=class_id)
    
    is_admin = ClassAdmin.objects.filter(class_obj=class_obj, user=request.user).exists()
    if not is_admin:
        messages.error(request, '只有班级管理员才能编辑班级信息')
        return redirect('class_detail', class_id=class_id)
    
    if request.method == 'POST':
        class_obj.name = request.POST.get('name', '').strip()
        class_obj.description = request.POST.get('description', '').strip()
        join_rule = request.POST.get('join_rule', 'approval').strip()
        class_obj.join_rule = join_rule
        class_obj.save()
        messages.success(request, '班级信息已更新')
        return redirect('class_detail', class_id=class_id)
    
    return render(request, 'quiz/frontend/edit_class.html', {'class_obj': class_obj})

@login_required
def delete_class(request, class_id):
    class_obj = get_object_or_404(Class, pk=class_id)
    
    is_admin = ClassAdmin.objects.filter(class_obj=class_obj, user=request.user).exists()
    if not is_admin:
        messages.error(request, '只有班级管理员才能删除班级')
        return redirect('class_detail', class_id=class_id)
    
    if request.method == 'POST':
        class_obj.delete()
        messages.success(request, '班级已删除')
        return redirect('class_list')
    
    return render(request, 'quiz/frontend/delete_class.html', {'class_obj': class_obj})

@login_required
def add_class_admin(request, class_id):
    class_obj = get_object_or_404(Class, pk=class_id)
    
    is_admin = ClassAdmin.objects.filter(class_obj=class_obj, user=request.user).exists()
    if not is_admin:
        messages.error(request, '只有班级管理员才能添加管理员')
        return redirect('class_detail', class_id=class_id)
    
    if request.method == 'POST':
        username = request.POST.get('username', '').strip()
        try:
            user = User.objects.get(username=username)
            if ClassAdmin.objects.filter(class_obj=class_obj, user=user).exists():
                messages.error(request, f'用户 {username} 已经是该班级的管理员')
            else:
                ClassAdmin.objects.create(class_obj=class_obj, user=user)
                messages.success(request, f'用户 {username} 已成为该班级的管理员')
        except User.DoesNotExist:
            messages.error(request, f'用户 {username} 不存在')
    
    return render(request, 'quiz/frontend/add_class_admin.html', {'class_obj': class_obj})

@login_required
def remove_class_admin(request, class_id, admin_id):
    class_obj = get_object_or_404(Class, pk=class_id)
    
    is_admin = ClassAdmin.objects.filter(class_obj=class_obj, user=request.user).exists()
    if not is_admin:
        messages.error(request, '只有班级管理员才能移除管理员')
        return redirect('class_detail', class_id=class_id)
    
    admin = get_object_or_404(ClassAdmin, pk=admin_id, class_obj=class_obj)
    
    if request.method == 'POST':
        if admin.user == request.user:
            messages.error(request, '不能移除自己')
            return redirect('class_detail', class_id=class_id)
        
        admin.delete()
        messages.success(request, '已移除该管理员')
        return redirect('class_detail', class_id=class_id)
    
    return render(request, 'quiz/frontend/remove_class_admin.html', {
        'class_obj': class_obj,
        'admin': admin
    })

@login_required
def assign_student(request, class_id):
    class_obj = get_object_or_404(Class, pk=class_id)
    
    is_admin = ClassAdmin.objects.filter(class_obj=class_obj, user=request.user).exists()
    if not is_admin:
        messages.error(request, '只有班级管理员才能分配学生')
        return redirect('class_detail', class_id=class_id)
    
    if request.method == 'POST':
        username = request.POST.get('username', '').strip()
        try:
            user = User.objects.get(username=username)
            
            if Class.objects.filter(id=class_id, profiles__user=user).exists():
                messages.error(request, f'用户 {username} 已经在该班级中')
            else:
                try:
                    profile = Profile.objects.get(user=user)
                    profile.class_obj = class_obj
                    profile.approval_status = 1
                    profile.save()
                except Profile.DoesNotExist:
                    Profile.objects.create(user=user, class_obj=class_obj, approval_status=1)
                
                ClassApplication.objects.filter(user=user, class_obj=class_obj).update(status=1)
                
                messages.success(request, f'用户 {username} 已添加到该班级')
        except User.DoesNotExist:
            messages.error(request, f'用户 {username} 不存在')
    
    return render(request, 'quiz/frontend/assign_student.html', {'class_obj': class_obj})

@login_required
def remove_student(request, class_id, user_id):
    class_obj = get_object_or_404(Class, pk=class_id)
    
    is_admin = ClassAdmin.objects.filter(class_obj=class_obj, user=request.user).exists()
    if not is_admin:
        messages.error(request, '只有班级管理员才能移除学生')
        return redirect('class_detail', class_id=class_id)
    
    student = get_object_or_404(User, pk=user_id)
    
    if request.method == 'POST':
        try:
            profile = Profile.objects.get(user=student)
            profile.class_obj = None
            profile.save()
        except Profile.DoesNotExist:
            pass
        
        messages.success(request, f'学生 {student.username} 已从班级移除')
        return redirect('class_detail', class_id=class_id)
    
    return render(request, 'quiz/frontend/remove_student.html', {
        'class_obj': class_obj,
        'student': student
    })

def apply_to_class(request):
    if request.method == 'POST':
        class_code = request.POST.get('class_code', '').strip()
        
        if not class_code:
            messages.error(request, '请输入班级代码')
            return render(request, 'quiz/frontend/apply_to_class.html')
        
        try:
            class_obj = Class.objects.get(code=class_code)
        except Class.DoesNotExist:
            messages.error(request, '班级代码不存在')
            return render(request, 'quiz/frontend/apply_to_class.html')
        
        if ClassApplication.objects.filter(user=request.user, class_obj=class_obj, status=0).exists():
            messages.error(request, '您已经申请过该班级，请等待审核')
            return render(request, 'quiz/frontend/apply_to_class.html')
        
        # 检查是否已经在班级中
        try:
            profile = Profile.objects.get(user=request.user)
            if profile.class_obj == class_obj:
                messages.error(request, '您已经在该班级中')
                return render(request, 'quiz/frontend/apply_to_class.html')
        except Profile.DoesNotExist:
            pass
        
        # 根据进班规则处理申请
        if class_obj.join_rule == 'auto':
            # 自动进班
            application = ClassApplication.objects.create(
                class_obj=class_obj,
                user=request.user,
                message=request.POST.get('message', ''),
                status=1
            )
            
            try:
                profile = Profile.objects.get(user=request.user)
                profile.class_obj = class_obj
                profile.approval_status = 1
                profile.save()
            except Profile.DoesNotExist:
                Profile.objects.create(user=request.user, class_obj=class_obj, approval_status=1)
            
            messages.success(request, f'已成功加入班级 "{class_obj.name}"！')
            return redirect('user_center')
        else:
            # 需要审核
            ClassApplication.objects.create(
                class_obj=class_obj,
                user=request.user,
                message=request.POST.get('message', '')
            )
            
            messages.success(request, f'已成功申请加入班级 "{class_obj.name}"，请等待管理员审核')
            return redirect('user_center')
    
    return render(request, 'quiz/frontend/apply_to_class.html')

@login_required
def my_applications(request):
    applications = ClassApplication.objects.filter(user=request.user).select_related('class_obj').order_by('-created_at')
    
    for app in applications:
        app.status_text = {0: '待审核', 1: '已通过', 2: '已拒绝'}.get(app.status, '未知')
    
    return render(request, 'quiz/frontend/my_applications.html', {
        'applications': applications
    })

@login_required
def class_applications(request, class_id):
    class_obj = get_object_or_404(Class, pk=class_id)
    
    is_admin = ClassAdmin.objects.filter(class_obj=class_obj, user=request.user).exists()
    if not is_admin:
        messages.error(request, '只有班级管理员才能查看申请列表')
        return redirect('class_detail', class_id=class_id)
    
    pending_applications = ClassApplication.objects.filter(class_obj=class_obj, status=0).select_related('user')
    processed_applications = ClassApplication.objects.filter(
        class_obj=class_obj,
        status__in=[1, 2]
    ).select_related('user', 'reviewed_by').order_by('-reviewed_at')[:20]
    
    return render(request, 'quiz/frontend/class_applications.html', {
        'class_obj': class_obj,
        'pending_applications': pending_applications,
        'processed_applications': processed_applications
    })

@login_required
def approve_application(request, class_id, application_id):
    class_obj = get_object_or_404(Class, pk=class_id)
    
    is_admin = ClassAdmin.objects.filter(class_obj=class_obj, user=request.user).exists()
    if not is_admin:
        messages.error(request, '只有班级管理员才能审批申请')
        return redirect('class_detail', class_id=class_id)
    
    application = get_object_or_404(ClassApplication, pk=application_id, class_obj=class_obj)
    
    application.status = 1
    application.reviewed_by = request.user
    application.save()
    
    try:
        profile = Profile.objects.get(user=application.user)
        profile.class_obj = class_obj
        profile.approval_status = 1
        profile.save()
    except Profile.DoesNotExist:
        Profile.objects.create(user=application.user, class_obj=class_obj, approval_status=1)
    
    messages.success(request, f'已批准 {application.user.username} 的加入申请')
    return redirect('class_applications', class_id=class_id)

@login_required
def reject_application(request, class_id, application_id):
    class_obj = get_object_or_404(Class, pk=class_id)
    
    is_admin = ClassAdmin.objects.filter(class_obj=class_obj, user=request.user).exists()
    if not is_admin:
        messages.error(request, '只有班级管理员才能审批申请')
        return redirect('class_detail', class_id=class_id)
    
    application = get_object_or_404(ClassApplication, pk=application_id, class_obj=class_obj)
    
    if request.method == 'POST':
        application.status = 2
        application.reviewed_by = request.user
        application.save()
        messages.success(request, f'已拒绝 {application.user.username} 的加入申请')
        return redirect('class_applications', class_id=class_id)
    
    return render(request, 'quiz/frontend/reject_application.html', {
        'class_obj': class_obj,
        'application': application
    })

@login_required
def class_assignments(request, class_id):
    class_obj = get_object_or_404(Class, pk=class_id)
    
    is_admin = ClassAdmin.objects.filter(class_obj=class_obj, user=request.user).exists()
    if not is_admin:
        messages.error(request, '只有班级管理员才能查看作业')
        return redirect('class_detail', class_id=class_id)
    
    assignments = ClassAssignment.objects.filter(class_obj=class_obj).order_by('-created_at')
    
    return render(request, 'quiz/frontend/class_assignments.html', {
        'class_obj': class_obj,
        'assignments': assignments
    })

@login_required
def create_class_assignment(request, class_id):
    class_obj = get_object_or_404(Class, pk=class_id)
    
    is_admin = ClassAdmin.objects.filter(class_obj=class_obj, user=request.user).exists()
    if not is_admin:
        messages.error(request, '只有班级管理员才能创建作业')
        return redirect('class_detail', class_id=class_id)
    
    # 获取已发布的试卷
    available_papers = TestPaper.objects.filter(is_published=True)
    
    if request.method == 'POST':
        title = request.POST.get('title', '').strip()
        description = request.POST.get('description', '').strip()
        assignment_type = request.POST.get('type', '1')
        paper_id = request.POST.get('paper_id')
        deadline = request.POST.get('deadline')
        time_limit = request.POST.get('time_limit')
        
        if not title:
            messages.error(request, '请填写作业标题')
            return render(request, 'quiz/frontend/create_class_assignment.html', {
                'class_obj': class_obj,
                'available_papers': available_papers
            })
        
        if not paper_id:
            messages.error(request, '请选择试卷')
            return render(request, 'quiz/frontend/create_class_assignment.html', {
                'class_obj': class_obj,
                'available_papers': available_papers
            })
        
        assignment = ClassAssignment.objects.create(
            class_obj=class_obj,
            title=title,
            description=description,
            type=int(assignment_type),
            deadline=parse_datetime_local(deadline) if deadline else None,
            time_limit=int(time_limit) if (time_limit and assignment_type == '2') else None,
            test_paper=TestPaper.objects.get(id=paper_id),
            is_allow_exam=True
        )
        
        messages.success(request, f'{"考试" if assignment_type == "2" else "作业"} "{title}" 创建成功！')
        return redirect('class_assignments', class_id=class_id)
    
    return render(request, 'quiz/frontend/create_class_assignment.html', {
        'class_obj': class_obj,
        'available_papers': available_papers
    })

@login_required
def class_assignment_detail(request, class_id, assignment_id):
    assignment = get_object_or_404(ClassAssignment, pk=assignment_id, class_obj_id=class_id)
    
    is_admin = ClassAdmin.objects.filter(class_obj=assignment.class_obj, user=request.user).exists()
    
    records = ClassAssignmentRecord.objects.filter(assignment=assignment).select_related('user')
    
    total_students = assignment.get_total_students()
    completed_count = assignment.get_completed_count()
    not_submitted_count = total_students - completed_count
    
    return render(request, 'quiz/frontend/class_assignment_detail.html', {
        'assignment': assignment,
        'class_obj': assignment.class_obj,
        'records': records,
        'is_admin': is_admin,
        'total_students': total_students,
        'completed_count': completed_count,
        'not_submitted_count': not_submitted_count
    })

@login_required
def publish_class_assignment(request, class_id, assignment_id):
    class_obj = get_object_or_404(Class, pk=class_id)
    
    is_admin = ClassAdmin.objects.filter(class_obj=class_obj, user=request.user).exists()
    if not is_admin:
        messages.error(request, '只有班级管理员才能发布作业')
        return redirect('class_detail', class_id=class_id)
    
    assignment = get_object_or_404(ClassAssignment, pk=assignment_id, class_obj=class_obj)
    
    if request.method == 'POST':
        assignment.status = 1
        assignment.published_at = timezone.now()
        assignment.save()
        
        students = Profile.objects.filter(class_obj=class_obj, approval_status=1)
        for student in students:
            ClassAssignmentRecord.objects.get_or_create(
                assignment=assignment,
                user=student.user
            )
        
        messages.success(request, f'作业 "{assignment.title}" 已发布！')
        return redirect('class_assignments', class_id=class_id)
    
    return render(request, 'quiz/frontend/publish_class_assignment.html', {
        'class_obj': class_obj,
        'assignment': assignment
    })


@login_required
def allow_exam(request, class_id, assignment_id):
    """允许考试"""
    class_obj = get_object_or_404(Class, pk=class_id)
    assignment = get_object_or_404(ClassAssignment, pk=assignment_id, class_obj=class_obj)
    
    is_admin = ClassAdmin.objects.filter(class_obj=class_obj, user=request.user).exists()
    if not is_admin:
        messages.error(request, '只有班级管理员才能操作')
        return redirect('class_assignment_detail', class_id=class_id, assignment_id=assignment_id)
    
    assignment.is_allow_exam = True
    assignment.save()
    
    messages.success(request, f'已允许考试：{assignment.title}')
    return redirect('class_assignment_detail', class_id=class_id, assignment_id=assignment_id)


@login_required
def close_exam(request, class_id, assignment_id):
    """关闭考试"""
    class_obj = get_object_or_404(Class, pk=class_id)
    assignment = get_object_or_404(ClassAssignment, pk=assignment_id, class_obj=class_obj)
    
    is_admin = ClassAdmin.objects.filter(class_obj=class_obj, user=request.user).exists()
    if not is_admin:
        messages.error(request, '只有班级管理员才能操作')
        return redirect('class_assignment_detail', class_id=class_id, assignment_id=assignment_id)
    
    assignment.is_allow_exam = False
    assignment.save()
    
    messages.success(request, f'已关闭考试：{assignment.title}')
    return redirect('class_assignment_detail', class_id=class_id, assignment_id=assignment_id)


@login_required
def reset_exam_status(request, class_id, assignment_id):
    """重置考试状态（清空所有答题记录）"""
    class_obj = get_object_or_404(Class, pk=class_id)
    assignment = get_object_or_404(ClassAssignment, pk=assignment_id, class_obj=class_obj)
    
    is_admin = ClassAdmin.objects.filter(class_obj=class_obj, user=request.user).exists()
    if not is_admin:
        messages.error(request, '只有班级管理员才能操作')
        return redirect('class_assignment_detail', class_id=class_id, assignment_id=assignment_id)
    
    if request.method == 'POST':
        ClassAssignmentRecord.objects.filter(assignment=assignment).delete()
        
        students = Profile.objects.filter(class_obj=class_obj, approval_status=1)
        for student in students:
            ClassAssignmentRecord.objects.create(assignment=assignment, user=student.user)
        
        messages.success(request, f'已重置考试状态：{assignment.title}')
        return redirect('class_assignment_detail', class_id=class_id, assignment_id=assignment_id)
    
    return render(request, 'quiz/frontend/reset_exam_status.html', {
        'class_obj': class_obj,
        'assignment': assignment
    })


@login_required
def student_class_assignments(request):
    try:
        profile = Profile.objects.get(user=request.user)
        if not profile.class_obj:
            messages.error(request, '您还没有加入任何班级')
            return redirect('user_center')
    except Profile.DoesNotExist:
        messages.error(request, '请先完善您的个人信息')
        return redirect('user_center')
    
    # 获取类型参数
    current_type = request.GET.get('type', '1')
    try:
        current_type = int(current_type)
    except:
        current_type = 1
    
    assignments = ClassAssignment.objects.filter(
        class_obj=profile.class_obj,
        status=1,
        type=current_type
    ).order_by('-published_at')
    
    records = ClassAssignmentRecord.objects.filter(user=request.user)
    record_dict = {r.assignment_id: r for r in records}
    
    assignment_list = []
    now = timezone.now()
    for assignment in assignments:
        record = record_dict.get(assignment.id)
        is_submitted = record and record.is_submitted
        is_overdue = assignment.deadline < now
        assignment_list.append({
            'assignment': assignment,
            'record': record,
            'is_submitted': is_submitted,
            'is_overdue': is_overdue
        })
    
    return render(request, 'quiz/frontend/student_class_assignments.html', {
        'class_obj': profile.class_obj,
        'assignment_list': assignment_list,
        'current_type': current_type,
        'now': now
    })

@login_required
def do_class_assignment(request, assignment_id):
    assignment = get_object_or_404(ClassAssignment, pk=assignment_id)
    
    if assignment.status != 1:
        messages.error(request, '该作业尚未发布')
        return redirect('student_class_assignments')
    
    # 考试模式：检查是否允许考试
    if assignment.type == 2 and not assignment.is_allow_exam:
        messages.error(request, '该考试已关闭')
        return redirect('student_class_assignments')
    
    # 获取用户最新的答题记录
    latest_record = ClassAssignmentRecord.objects.filter(
        assignment=assignment,
        user=request.user
    ).order_by('-attempt').first()
    
    # 考试模式：只能有一次提交
    if assignment.type == 2 and latest_record and latest_record.is_submitted:
        messages.error(request, '您已经提交过该考试')
        return redirect('student_class_assignments')
    
    # 检查考试时间限制
    if assignment.type == 2 and assignment.time_limit:
        if latest_record and latest_record.start_time:
            # 检查是否超时
            time_elapsed = (timezone.now() - latest_record.start_time).total_seconds() / 60
            if time_elapsed > assignment.time_limit:
                messages.error(request, '考试已超时，自动提交')
                # 自动提交当前答案（如果有）
                return redirect('student_class_assignments')
    
    # 创建新的答题记录（作业模式允许多次）
    if not latest_record or latest_record.is_submitted:
        attempt = latest_record.attempt + 1 if latest_record else 1
        record = ClassAssignmentRecord.objects.create(
            assignment=assignment,
            user=request.user,
            start_time=timezone.now(),
            attempt=attempt
        )
    else:
        record = latest_record
        if not record.start_time:
            record.start_time = timezone.now()
            record.save()
    
    if request.method == 'POST':
        if assignment.test_paper:
            test_paper = assignment.test_paper
            questions = list(test_paper.questions.all())
            
            user_answers = {}
            for q in questions:
                answer_key = f'question_{q.id}'
                if answer_key in request.POST:
                    user_answers[q.id] = request.POST[answer_key]
            
            score, correct_count, wrong_count, total_count, question_results = calculate_score(questions, user_answers)
            
            record.score = score
            record.is_submitted = True
            record.submitted_at = timezone.now()
            record.save()
            
            test_record = TestRecord.objects.create(
                user=request.user,
                test_paper=test_paper,
                score=score,
                total_score=test_paper.total_score,
                completed_at=timezone.now()
            )
            
            for result in question_results:
                AnswerRecord.objects.create(
                    test_record=test_record,
                    question=result['question'],
                    user_answer=result.get('user_answer', ''),
                    correct_answer=result['correct_answer'],
                    is_correct=result['is_correct'],
                    score=result['score']
                )
            
            messages.success(request, f'{"考试" if assignment.type == 2 else "作业"}提交成功！得分：{score}分')
        else:
            record.is_submitted = True
            record.submitted_at = timezone.now()
            record.save()
            messages.success(request, f'{"考试" if assignment.type == 2 else "作业"}提交成功！')
        
        return redirect('student_class_assignments')
    
    questions = []
    if assignment.test_paper:
        test_paper = assignment.test_paper
        questions = list(test_paper.questions.all())
        for q in questions:
            if isinstance(q.options, str):
                import json
                try:
                    q.options = json.loads(q.options)
                except:
                    q.options = {}
    
    # 计算剩余时间
    remaining_seconds = None
    if assignment.type == 2 and assignment.time_limit:
        if record.start_time:
            elapsed_seconds = (timezone.now() - record.start_time).total_seconds()
            remaining_seconds = int(max(0, assignment.time_limit * 60 - elapsed_seconds))
    
    # 作业模式：始终可以查看答案
    show_answer = assignment.type == 1
    
    return render(request, 'quiz/frontend/do_class_assignment.html', {
        'assignment': assignment,
        'questions': questions,
        'test_paper': test_paper,
        'remaining_seconds': remaining_seconds,
        'show_answer': show_answer,
        'record': record
    })

@login_required
def submit_class_assignment(request, assignment_id):
    assignment = get_object_or_404(ClassAssignment, pk=assignment_id)
    
    if request.method == 'POST':
        # 获取用户最新的答题记录
        latest_record = ClassAssignmentRecord.objects.filter(
            assignment=assignment,
            user=request.user
        ).order_by('-attempt').first()
        
        # 考试模式：只能提交一次
        if assignment.type == 2 and latest_record and latest_record.is_submitted:
            return JsonResponse({'success': False, 'message': '已经提交过'})
        
        # 使用当前记录或创建新记录
        if latest_record and not latest_record.is_submitted:
            record = latest_record
        else:
            # 创建新的答题记录（作业模式允许多次练习）
            attempt = latest_record.attempt + 1 if latest_record else 1
            record = ClassAssignmentRecord.objects.create(
                assignment=assignment,
                user=request.user,
                attempt=attempt
            )
        
        if assignment.test_paper:
            test_paper = assignment.test_paper
            questions = list(test_paper.questions.all())
            
            user_answers = {}
            for q in questions:
                answer_key = f'question_{q.id}'
                if q.type == 2:
                    # 多选题：获取所有选中的选项
                    selected_options = []
                    for opt in ['A', 'B', 'C', 'D']:
                        if f'question_{q.id}_{opt}' in request.POST:
                            selected_options.append(opt)
                    if selected_options:
                        user_answers[q.id] = ''.join(sorted(selected_options))
                elif answer_key in request.POST:
                    user_answers[q.id] = request.POST[answer_key]
            
            score, correct_count, wrong_count, total_count, question_results = calculate_score(questions, user_answers)
            
            record.score = score
            record.is_submitted = True
            record.submitted_at = timezone.now()
            record.save()
            
            test_record = TestRecord.objects.create(
                user=request.user,
                test_paper=test_paper,
                score=score,
                total_score=test_paper.total_score,
                completed_at=timezone.now()
            )
            
            for result in question_results:
                AnswerRecord.objects.create(
                    test_record=test_record,
                    question=result['question'],
                    user_answer=result.get('user_answer', ''),
                    correct_answer=result['correct_answer'],
                    is_correct=result['is_correct'],
                    score=result['score']
                )
            
            return JsonResponse({
                'success': True,
                'message': f'提交成功！得分：{score}分',
                'score': score
            })
        else:
            record.is_submitted = True
            record.submitted_at = timezone.now()
            record.save()
            return JsonResponse({'success': True, 'message': '提交成功！'})
    
    return JsonResponse({'success': False, 'message': '无效的请求'})

# 后台管理视图
@staff_member_required
def admin_import_questions(request):
    """后台导入试题 - 使用统一的导入函数，支持防重复导入"""
    if request.method == 'POST':
        if 'file' in request.FILES:
            file = request.FILES['file']
            
            file_content = file.read()
            file.seek(0)
            
            is_dup, prev_count, import_time = is_duplicate_import(file_content)
            if is_dup:
                messages.error(request, 
                    f'检测到重复导入！该文件已于 {import_time.strftime("%Y-%m-%d %H:%M")} 导入，'
                    f'共导入 {prev_count} 道题目。如需重新导入，请等待24小时或修改文件内容后重试。')
                return render(request, 'quiz/admin/import_questions.html', {'step': 1})
            
            questions_data, stats, errors = import_questions_from_excel(file)
            
            if errors:
                messages.error(request, errors[0])
                return render(request, 'quiz/admin/import_questions.html', {'step': 1})
            
            for idx, q in enumerate(questions_data):
                q['row'] = idx + 2
                q['has_error'] = not (q.get('correct_answer') and q.get('score'))
            
            request.session['import_file_hash'] = generate_file_hash(file_content)
            request.session['import_questions_data'] = questions_data
            
            return render(request, 'quiz/admin/import_questions.html', {
                'step': 2,
                'questions_data': questions_data,
                'questions_json': json.dumps(questions_data, ensure_ascii=False),
                'total_score': stats['total_score'],
                'valid_count': stats['valid_count'],
                'missing_count': stats['missing_count'],
                'errors': stats['errors']
            })
        
        elif 'questions_json' in request.POST:
            try:
                questions_data = json.loads(request.POST['questions_json'])
                imported_count = 0
                
                for q_data in questions_data:
                    if q_data.get('content') and q_data.get('correct_answer') and q_data.get('score'):
                        subject_obj = None
                        chapter_obj = None
                        section_obj = None
                        kp_objects = []
                        
                        subject_name = q_data.get('subject_name', '').strip()
                        chapter_title = q_data.get('chapter_title', '').strip()
                        section_title = q_data.get('section_title', '').strip()
                        kp_names = q_data.get('knowledge_points_str', '').strip()
                        
                        if subject_name:
                            subject_obj, _ = Subject.objects.get_or_create(
                                name=subject_name,
                                defaults={'code': subject_name[:10].upper(), 'icon': '📚'}
                            )
                        
                        if subject_obj and chapter_title:
                            chapter_obj, _ = Chapter.objects.get_or_create(
                                subject=subject_obj,
                                title=chapter_title,
                                defaults={'number': Chapter.objects.filter(subject=subject_obj).count() + 1}
                            )
                        
                        if chapter_obj and section_title:
                            section_obj, _ = Section.objects.get_or_create(
                                chapter=chapter_obj,
                                title=section_title,
                                defaults={'number': Section.objects.filter(chapter=chapter_obj).count() + 1}
                            )
                        
                        if kp_names and subject_obj:
                            for kp_name in kp_names.split(','):
                                kp_name = kp_name.strip()
                                if kp_name:
                                    kp, created = KnowledgePoint.objects.get_or_create(
                                        subject=subject_obj,
                                        name=kp_name,
                                        defaults={'section': section_obj}
                                    )
                                    kp_objects.append(kp)
                        
                        question = Question.objects.create(
                            type=q_data['type'],
                            content=q_data['content'],
                            options=q_data.get('options', {}),
                            correct_answer=q_data['correct_answer'],
                            score=q_data['score'],
                            explanation=q_data.get('explanation', ''),
                            subject=subject_obj,
                            chapter=chapter_obj,
                            section=section_obj,
                            is_public=True,
                            created_by='admin'
                        )
                        
                        if kp_objects:
                            question.knowledge_points.set(kp_objects)
                        
                        imported_count += 1
                
                if 'import_file_hash' in request.session:
                    file_hash = request.session['import_file_hash']
                    _imported_files_cache[file_hash] = (timezone.now(), imported_count)
                    del request.session['import_file_hash']
                    del request.session['import_questions_data']
                
                return render(request, 'quiz/admin/import_questions.html', {
                    'step': 3,
                    'imported_count': imported_count
                })
            
            except Exception as e:
                messages.error(request, f'导入失败：{str(e)}')
                return render(request, 'quiz/admin/import_questions.html', {'step': 1})
    
    return render(request, 'quiz/admin/import_questions.html', {'step': 1})

from django.http import JsonResponse

@staff_member_required
def admin_export_template(request):
    """下载后台导入模板"""
    return download_template_response()

def download_import_template(request):
    """下载前台导入模板"""
    return download_template_response()

@staff_member_required
def admin_create_testpaper(request):
    """后台组卷"""
    all_questions = Question.objects.all()
    
    if request.method == 'POST':
        title = request.POST.get('title')
        description = request.POST.get('description')
        is_published = request.POST.get('is_published') == 'on'
        selected_questions = request.POST.get('selected_questions')
        
        if not title:
            messages.error(request, '请输入试卷标题')
            return redirect('admin_create_testpaper')
        
        if not selected_questions:
            messages.error(request, '请选择至少一道题目')
            return redirect('admin_create_testpaper')
        
        test_paper = TestPaper.objects.create(
            title=title,
            description=description,
            created_by='admin',
            is_published=is_published
        )
        
        total_score = 0
        for q_id in selected_questions.split(','):
            try:
                question = Question.objects.get(id=int(q_id))
                test_paper.questions.add(question)
                total_score += question.score
            except:
                pass
        
        test_paper.total_score = total_score
        test_paper.save()
        
        messages.success(request, f'试卷 "{title}" 创建成功！')
        return redirect('admin_preview_testpaper', paper_id=test_paper.id)
    
    questions_list = []
    for q in all_questions:
        if isinstance(q.options, str):
            import json
            try:
                q.options = json.loads(q.options)
            except:
                q.options = {}
        questions_list.append(q)
    
    return render(request, 'quiz/admin/create_testpaper.html', {
        'all_questions': questions_list
    })

@staff_member_required
def admin_preview_testpaper(request, paper_id):
    test_paper = get_object_or_404(TestPaper, pk=paper_id)
    questions = list(test_paper.questions.all())
    
    for q in questions:
        if isinstance(q.options, str):
            import json
            try:
                q.options = json.loads(q.options)
            except:
                q.options = {}
    
    return render(request, 'quiz/admin/preview_testpaper.html', {
        'test_paper': test_paper,
        'questions': questions
    })

@staff_member_required
def admin_import_testpaper(request):
    """后台导入试卷"""
    if request.method == 'POST':
        file = request.FILES.get('file')
        title = request.POST.get('title', '导入试卷')
        description = request.POST.get('description', '')
        is_published = request.POST.get('is_published') == 'on'
        
        if not file:
            messages.error(request, '请上传文件')
            return render(request, 'quiz/admin/import_testpaper.html')
        
        try:
            questions_data, stats, errors = import_questions_from_excel(file)
            if errors:
                messages.error(request, errors[0])
                return render(request, 'quiz/admin/import_testpaper.html')
            
            test_paper = TestPaper.objects.create(
                title=title,
                description=description,
                created_by='admin',
                is_published=is_published
            )
            
            total_score = 0
            for q_data in questions_data:
                if not q_data.get('content') or not q_data.get('correct_answer'):
                    continue
                
                options_data = q_data.get('options', {})
                if isinstance(options_data, str):
                    try:
                        options_data = json.loads(options_data)
                    except:
                        options_data = {}
                
                q_type = int(q_data.get('type', 1))
                if q_type not in [1, 2, 3]:
                    q_type = 1
                
                q_score = q_data.get('score', 1)
                try:
                    q_score = int(q_score) if q_score else 1
                except:
                    q_score = 1
                
                subject_obj = None
                chapter_obj = None
                section_obj = None
                kp_objects = []
                
                subject_name = q_data.get('subject_name', '').strip()
                chapter_title = q_data.get('chapter_title', '').strip()
                section_title = q_data.get('section_title', '').strip()
                kp_names = q_data.get('knowledge_points_str', '').strip()
                
                if subject_name:
                    subject_obj, _ = Subject.objects.get_or_create(
                        name=subject_name,
                        defaults={'code': subject_name[:10].upper(), 'icon': '📚'}
                    )
                
                if subject_obj and chapter_title:
                    chapter_obj, _ = Chapter.objects.get_or_create(
                        subject=subject_obj,
                        title=chapter_title,
                        defaults={'number': Chapter.objects.filter(subject=subject_obj).count() + 1}
                    )
                
                if chapter_obj and section_title:
                    section_obj, _ = Section.objects.get_or_create(
                        chapter=chapter_obj,
                        title=section_title,
                        defaults={'number': Section.objects.filter(chapter=chapter_obj).count() + 1}
                    )
                
                if kp_names and subject_obj:
                    for kp_name in kp_names.split(','):
                        kp_name = kp_name.strip()
                        if kp_name:
                            kp, created = KnowledgePoint.objects.get_or_create(
                                subject=subject_obj,
                                name=kp_name,
                                defaults={'section': section_obj}
                            )
                            kp_objects.append(kp)
                
                question = Question.objects.create(
                    type=q_type,
                    content=q_data['content'],
                    options=options_data,
                    correct_answer=q_data['correct_answer'],
                    score=q_score,
                    explanation=q_data.get('explanation', ''),
                    subject=subject_obj,
                    chapter=chapter_obj,
                    section=section_obj,
                    is_public=True,
                    created_by='admin'
                )
                
                if kp_objects:
                    question.knowledge_points.set(kp_objects)
                
                test_paper.questions.add(question)
                total_score += q_score
            
            test_paper.total_score = total_score
            test_paper.save()
            
            messages.success(request, f'试卷 "{title}" 导入成功！共导入 {len(questions_data)} 道题目')
            return redirect('admin_preview_testpaper', paper_id=test_paper.id)
        except Exception as e:
            messages.error(request, f'导入失败：{str(e)}')
    
    return render(request, 'quiz/admin/import_testpaper.html')
