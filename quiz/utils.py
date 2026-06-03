from django.core.paginator import Paginator, PageNotAnInteger, EmptyPage
from django.http import HttpResponse
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils.exceptions import InvalidFileException
import json

def parse_options(options_str):
    """解析选项字段，如果是字符串则尝试JSON解析"""
    if isinstance(options_str, str):
        try:
            return json.loads(options_str)
        except (json.JSONDecodeError, ValueError):
            return {}
    return options_str or {}

def paginate_queryset(queryset, page_num, items_per_page=9):
    paginator = Paginator(queryset, items_per_page)
    try:
        paginated_items = paginator.page(page_num)
    except PageNotAnInteger:
        paginated_items = paginator.page(1)
    except EmptyPage:
        paginated_items = paginator.page(paginator.num_pages)
    return paginated_items

# 判断题答案映射，将各种表示统一为标准值
TRUE_VALUES = {'对', '正确', '是', 't', 'true', '1', 'yes', 'y'}
FALSE_VALUES = {'错', '错误', '否', 'f', 'false', '0', 'no', 'n'}

def normalize_judge_answer(answer):
    """标准化判断题答案"""
    if answer is None:
        return None
    answer = str(answer).strip().lower()
    if answer in TRUE_VALUES:
        return 'true'
    elif answer in FALSE_VALUES:
        return 'false'
    return answer

def compare_answers(user_answer, correct_answer):
    if user_answer is None:
        return False
    
    # 尝试判断题特殊处理
    normalized_user = normalize_judge_answer(user_answer)
    normalized_correct = normalize_judge_answer(correct_answer)
    
    # 如果是判断题答案（已标准化），使用标准化后的值比较
    if normalized_user in ('true', 'false') and normalized_correct in ('true', 'false'):
        return normalized_user == normalized_correct
    
    # 默认使用精确匹配
    return user_answer.strip().lower() == correct_answer.strip().lower()

def calculate_score(questions, user_answers):
    score = 0
    correct_count = 0
    question_results = []

    for question in questions:
        user_answer = user_answers.get(str(question.id), user_answers.get(question.id))
        is_correct = compare_answers(user_answer, question.correct_answer)

        if is_correct:
            score += question.score
            correct_count += 1
            result = '正确'
        elif user_answer is None:
            result = '未答'
        else:
            result = '错误'

        question_results.append({
            'question': question,
            'user_answer': user_answer,
            'correct_answer': question.correct_answer,
            'result': result,
            'score': question.score,
            'is_correct': is_correct
        })

    total_count = len(question_results)
    wrong_count = total_count - correct_count

    return score, correct_count, wrong_count, total_count, question_results

def parse_datetime_local(datetime_str):
    import datetime
    datetime_clean = datetime_str.replace('T', ' ')
    return datetime.datetime.strptime(datetime_clean, '%Y-%m-%d %H:%M')

class ExcelImporter:
    HEADER_ALIASES = {
        'content': ['content', '题目', '题目内容', 'question', '题干'],
        'type': ['type', '题型', '题目类型', '类别'],
        'option_a': ['option_a', '选项A', 'A', '选项a'],
        'option_b': ['option_b', '选项B', 'B', '选项b'],
        'option_c': ['option_c', '选项C', 'C', '选项c'],
        'option_d': ['option_d', '选项D', 'D', '选项d'],
        'options': ['options', '选项', '所有选项'],
        'correct_answer': ['correct_answer', '答案', '正确答案', '参考答案', 'answer'],
        'score': ['score', '分值', '分数', '得分'],
        'explanation': ['explanation', '解析', '答案解析', '解析说明'],
        'subject': ['subject', '学科', '科目', '所属学科', '学科名称'],
        'chapter': ['chapter', '章节', '所属章节', '章', '章节名称'],
        'section': ['section', '小节', '所属小节', '节', '小节名称'],
        'knowledge_points': ['knowledge_points', '知识点', '知识要点', '关联知识点', 'kp'],
    }

    TYPE_MAPPING = {
        '3': 3, '判断题': 3, '判断': 3, 'judge': 3,
        '2': 2, '多选题': 2, '多选': 2, 'multiple': 2,
        '1': 1, '单选题': 1, '单选': 1, 'single': 1, '选择题': 1, '选择': 1, 'choice': 1
    }

    def __init__(self, worksheet, required_keys=None):
        self.ws = worksheet
        self.errors = []
        self.warnings = []
        self.header_map = {}
        if required_keys is None:
            required_keys = ['content', 'correct_answer', 'score']
        self.required_keys = required_keys

    def parse_headers(self):
        headers = [cell.value for cell in self.ws[1]]
        for idx, header in enumerate(headers):
            if header:
                header_str = str(header).strip()
                header_lower = header_str.lower()
                for key, aliases in self.HEADER_ALIASES.items():
                    if header_lower in aliases or header_str in aliases:
                        self.header_map[key] = idx
                        break
        missing = [k for k in self.required_keys if k not in self.header_map]
        return missing

    def parse_row(self, row, row_idx):
        try:
            content = str(row[self.header_map['content']].value or '').strip()
            if not content:
                return None

            q_type = 1
            if 'type' in self.header_map:
                type_val = row[self.header_map['type']].value
                if type_val is not None:
                    type_str = str(type_val).strip()
                    q_type = self.TYPE_MAPPING.get(type_str, 1)

            options = self._parse_options(row)
            correct_answer = str(row[self.header_map['correct_answer']].value or '').strip()
            score = self._parse_score(row)
            explanation = ''
            if 'explanation' in self.header_map:
                explanation = str(row[self.header_map['explanation']].value or '').strip()

            subject_name = ''
            if 'subject' in self.header_map:
                subject_val = row[self.header_map['subject']].value
                if subject_val:
                    subject_name = str(subject_val).strip()

            chapter_title = ''
            if 'chapter' in self.header_map:
                chapter_val = row[self.header_map['chapter']].value
                if chapter_val:
                    chapter_title = str(chapter_val).strip()

            section_title = ''
            if 'section' in self.header_map:
                section_val = row[self.header_map['section']].value
                if section_val:
                    section_title = str(section_val).strip()

            knowledge_points_str = ''
            if 'knowledge_points' in self.header_map:
                kp_val = row[self.header_map['knowledge_points']].value
                if kp_val:
                    knowledge_points_str = str(kp_val).strip()

            has_error = not correct_answer or (not score and score != 0)
            if not correct_answer:
                self.errors.append(f'第{row_idx}行：正确答案为空')
            if not score and score != 0:
                self.errors.append(f'第{row_idx}行：分值格式错误')

            return {
                'content': content,
                'type': q_type,
                'options': options,
                'correct_answer': correct_answer,
                'score': score,
                'explanation': explanation,
                'subject_name': subject_name,
                'chapter_title': chapter_title,
                'section_title': section_title,
                'knowledge_points_str': knowledge_points_str,
                'row': row_idx,
                'has_error': has_error
            }
        except Exception as e:
            self.errors.append(f'第{row_idx}行：{str(e)}')
            return None

    def _parse_options(self, row):
        options = {}
        option_cols = ['option_a', 'option_b', 'option_c', 'option_d']
        option_letters = ['A', 'B', 'C', 'D']

        for col_key, letter in zip(option_cols, option_letters):
            if col_key in self.header_map:
                val = row[self.header_map[col_key]].value
                if val and str(val).strip():
                    options[letter] = str(val).strip()

        if not options and 'options' in self.header_map:
            options_str = str(row[self.header_map['options']].value or '').strip()
            if options_str:
                for item in options_str.split(','):
                    item = item.strip()
                    if item and len(item) >= 2:
                        letter = item[0].upper()
                        if letter in ['A', 'B', 'C', 'D']:
                            options[letter] = item[1:].strip()

        return options

    def _parse_score(self, row):
        try:
            score = int(row[self.header_map['score']].value or 0)
            return score if score > 0 else ''
        except:
            return ''

    def parse_all(self):
        questions_data = []
        for row_idx, row in enumerate(self.ws.iter_rows(min_row=2), start=2):
            parsed = self.parse_row(row, row_idx)
            if parsed:
                questions_data.append(parsed)
        return questions_data

def create_import_template():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "题目导入模板"

    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="667EEA", end_color="764BA2", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )

    headers = ['题目内容', '题型', '选项A', '选项B', '选项C', '选项D', '正确答案', '分值', '解析', '学科', '章节', '小节', '知识点']
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border

    example_data = [
        ['以下哪个是Python的关键字？', '单选题', 'and', 'or', 'true', 'false', 'A', 5, 'and是Python的关键字', '计算机', '第一章 基础语法', '1.1 关键字', '关键字,标识符'],
        ['下列哪些是Python的数据类型？', '多选题', 'int', 'str', 'list', 'dict', 'ABCD', 10, 'Python支持多种数据类型', '计算机', '第一章 基础语法', '1.2 数据类型', '数据类型,int,str'],
        ['Python是一种编程语言', '判断题', '', '', '', '', '正确', 3, 'Python确实是编程语言', '计算机', '第一章 基础语法', '1.1 关键字', '编程语言'],
    ]

    for row_idx, row_data in enumerate(example_data, start=2):
        for col_idx, value in enumerate(row_data, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.alignment = Alignment(vertical="center", wrap_text=True)
            cell.border = thin_border

    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 8
    ws.column_dimensions['C'].width = 12
    ws.column_dimensions['D'].width = 12
    ws.column_dimensions['E'].width = 12
    ws.column_dimensions['F'].width = 12
    ws.column_dimensions['G'].width = 10
    ws.column_dimensions['H'].width = 8
    ws.column_dimensions['I'].width = 20
    ws.column_dimensions['J'].width = 12
    ws.column_dimensions['K'].width = 18
    ws.column_dimensions['L'].width = 15
    ws.column_dimensions['M'].width = 25
    ws.freeze_panes = 'A2'

    return wb

def download_template_response(filename='题目导入模板.xlsx'):
    wb = create_import_template()
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename={filename}'
    wb.save(response)
    return response

def import_questions_from_excel(file, subject_map=None, chapter_map=None, section_map=None, knowledge_point_map=None):
    try:
        wb = openpyxl.load_workbook(file)
        ws = wb.active
        importer = ExcelImporter(ws)
        missing = importer.parse_headers()
        if missing:
            return None, None, [f'缺少必需列：{", ".join(missing)}']
        questions_data = importer.parse_all()
        if not questions_data:
            return None, None, importer.errors[:5] if importer.errors else ['文件中没有有效的题目数据']
        valid_count = sum(1 for q in questions_data if q.get('correct_answer') and q.get('score'))
        missing_count = len(questions_data) - valid_count
        total_score = sum(q['score'] if isinstance(q['score'], int) else 0 for q in questions_data)
        return questions_data, {
            'total_score': total_score,
            'valid_count': valid_count,
            'missing_count': missing_count,
            'errors': importer.errors[:10]
        }, None
    except InvalidFileException:
        return None, None, ['文件格式不正确，请上传 .xlsx 格式的 Excel 文件']
    except Exception as e:
        return None, None, [f'读取文件失败：{str(e)}']
